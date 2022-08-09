from unicodedata import decimal
import pandas as pd
import numpy as np
import datetime
#- https://genchan.net/it/programming/python/4949/#i-3
from functions import module_01
#- 参照を2行目から最終行まで行う列を色付けするために導入 - 20220405
import openpyxl
from openpyxl.styles import PatternFill
import os
import glob

#- 20220530
from pathlib import Path

#! I added the below time related things
t_delta = datetime.timedelta(hours=9)
JST = datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)
# print(repr(now))

today = now.strftime('%Y%m%d')

d = now.strftime('%Y%m%d%H%M%S')
# print(d)  #example 20211104173728 秒まで

d = now.strftime('%Y/%m/%d %H:%M')
# print(d)  #example 2021/11/04 17:37 分まで

#- 20220530
cwd = Path.cwd()

def neutralize_path_str(path):
    #- パスを繋げ、パスオブジェクトから文字列に変換する
    new_path = str(cwd / path)
    #- パス同士をつなげ完成したファイルのディレクトリの区切り文字を / に変更
    return new_path.replace(os.sep,'/')

def neutralize_path_obj(path):
    #- パスを繋げ、パスオブジェクトから文字列に変換する
    new_path = cwd / path
    #- パス同士をつなげ完成したファイルのディレクトリの区切り文字を / に変更
    print('\n', 'new_path_obj: ',new_path, '\n')
    #^ TypeError: replace() takes 2 positional arguments but 3 were given
    #! 不可 オブジェクトにreplace()を使おうとすると上記のエラーが発生する
    # return new_path.replace(os.sep,'/')
    return new_path

# words_to_replace = pd.read_csv('csv_folder/csv_replace_words/words_to_replace.csv')
#- csvファイルからexcelファイルに変更 - 20220416
#- 20220530 - macとwindows両方で互換性のあるpathにする
path_str = neutralize_path_str('csv_folder/csv_replace_words')
words_to_replace_separate = pd.read_excel(f'{path_str}/words_to_replace.xlsx', sheet_name='words_to_replace_separate')

path_to_ebay_str = neutralize_path_str('csv_folder/csv_for_concat/ebay')
searched_file_loc  = f'{path_to_ebay_str}/main_data_{today}.xlsx'

print('\n', 'searched_file_loc: ', searched_file_loc ,'\n')

df = pd.read_excel(searched_file_loc)

# print('\n','words_to_replace_separate: ', words_to_replace_separate)

# load the dataframe to dictionary for easier processing
words_dict = {}
for index, rows in words_to_replace_separate.iterrows():
    if rows["new word"] is not np.nan:
        key = rows["old word"]
        value = rows["new word"]
    else:
        key = rows["old word"]
        value = " "
    words_dict[key] = value
words_dict

# convert to proper datatypes
df = df.convert_dtypes()

# print('\n','入替え用のExcelファイルを辞書型に変更: ',words_dict)

#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
# 容れ物
updated_old = ''

# main function to replace old values with new ones
def helper_function(old):
    # updated_old = ''
    if not isinstance(old, str):
        return old
    # old = old.split(' ')
    # print('\n','old: ',old)

    #* キーと値をセットでループ処理する場合は、.items() メソッドを使用
    # https://blog.kikagaku.co.jp/python-for-dictionary
    for key, value in words_dict.items():
        # print('key: ',key)
        # print('value: ',value)
        while key in old:
            # print(f'Changed {old} to {words_dict[key]}, {key}')
            # old[i] = words_dict[key]
            old = old.replace(key,value).strip()
            updated_old = old
            # print('updated_old',updated_old)
            # if key == '汚れ有':
            #     print('break','='*30)
            #     break
        updated_old = old
    return updated_old


#- 置き換えの対象を特定の列名に絞り込んだ-20220328 df[['*Title','U-title','X-category']]:
#- 20220411 shopify用
for index, rows in df.iterrows():
    for each_col in df[['U-title']]:
        new_val = helper_function(df[each_col][index])
        df[each_col][index] = new_val

# print('df-2: ', df)

OUTPUT_FILENAME_Class = module_01.Concat_file_names() #*インスタンス化

print('\n', '確認01-102'  ,'\n')

#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#- 20220530 pathlibで相対パスを作成
path_to_ebay_obj = Path('csv_folder/csv_for_concat/ebay')

excel_files = list(path_to_ebay_obj.glob('*.xlsx'))
#! [WindowsPath('csv_folder/csv_for_concat/ebay/2nd-Alexander McQueen-sneaker-Men-202205191547.xlsx'), WindowsPath('csv_folder/csv_for_concat/ebay/ebay-TESTsneaker--Men-202205301603.xlsx'), WindowsPath('csv_folder/csv_for_concat/ebay/main_data_20220530.xlsx')] となっており、リストの要素をstring型にしないと使用できない
excel_files = [str(excel_file) for excel_file in excel_files ]
print(excel_files)

#- 条件分岐で仕入先が2ndStでもtrefacでも対応できるようにするための条件付に利用
#- xlsxのファイル名を全て文字列でつなげる
excel_files_str = '_'.join(excel_files)
print('\n', 'excel_files_str: ',excel_files_str, '\n')



#* クラス内のメソッドにconcatするファイルの場所を渡す
#- 2つ目の引数で仕入元を入力する 例: 2nd, trf, vec - 20220424
if 'trf' in excel_files_str:
    OUTPUT_FILENAME = OUTPUT_FILENAME_Class.concat_file_names(f'{path_to_ebay_str}/','trf' )

if '2nd' in excel_files_str:
    OUTPUT_FILENAME = OUTPUT_FILENAME_Class.concat_file_names(f'{path_to_ebay_str}/','2nd' )

# OUTPUT_FILENAME = '00-W-bag-20220405.xlsx'
print('='*40)
print('\n','outputName: ',OUTPUT_FILENAME)
print('='*40)

#- 20220510
sheet_name_01 = 'reference'
sheet_name_02 = 'ref-copy' #正直不要

#-https://obgynai.com/python-pandas-excel-export/
#* 一つのExcelファイルにシートを複数作成するための処理 - 20220330
with pd.ExcelWriter(OUTPUT_FILENAME) as writer:
    #- ^を置換えるにはエスケープが必要 replaceの部分一致置き換え方法 - 20220411
    #! https://it-ojisan.tokyo/pandas-replace-regex/
    df = df.replace('\^','',regex=True)
    df.to_excel(writer, sheet_name=sheet_name_01, index=False) #シート='main'
    #- 使用しないが元のクラス内のメソッドでエラーが出ないように作成
    df = df.replace('\^','',regex=True)
    df.to_excel(writer, sheet_name=sheet_name_02, index=False) #シート='reference'

#- module01のクラス内のchange_excel_columns_color関数に渡す上記エクセルファイルの場所-20220510
#- 20220530 現在いるフォルダを指定
load_workbook_location = str(cwd)

print('\n', 'OUTPUT_FILENAME: ',OUTPUT_FILENAME, '\n')

result_file_location = neutralize_path_obj('results/ebay')

ChangeExcelColor_CLASS = module_01.ChangeExcelColorRevised(load_workbook_location,OUTPUT_FILENAME,result_file_location)

#- 色分けしたい種類分、辞書型で変数を作成する 青、赤、緑なら３つ用意する
# ff00ff
pink_columns =  {'columns':['Ref-title_category_combined','Ref-translation','URef-title-completed','Ref-len()','Ref-material_translation','X-size-tr','Xd-brand','Xd-item','Xd-staff-comment','Xd-color','Xd-line','Xd-model','Xd-country','Xd-remark','Xd-material','Xd-brand-tr','Xd-item-tr','Xd-staff-comment-tr','Xd-color-tr','Xd-line-tr'],'color':'ff00ff'}

# f7b977
orange_columns =  {'columns':['PicURL','CustomLabel','*Description','ConditionDescription','*StartPrice','MinimumBestOfferPrice',],'color':'f7b977'}

# 5bf6ff
lightBlue_columns =  {'columns':['U-title','*Title','*Category','*C:US Shoe Size','C:Movement'],'color':'5bf6ff'}

# 6fa7ff
#- '要確認-M-W'はwWatch限定 - womenかmenかを見極める category#が同じだが、departmentを変更するため- 20220617
blue_columns =  {'columns':['C:Brand','X-model','X-condition','X-condition_detail','要確認-M-W','X-tag-size'],'color':'6fa7ff'}

# ffe999
lightYellow_columns =  {
    'columns':['URef-size_for_ConditionDescription','ConditionID','ShippingProfileName'],
    'color':'ffe999'
}

yellow_columns =  {
    'columns': ['costPrice + s/p','url','costPrice + s/p'],
    'color':'fff700'
}

#* 色分けしたい種類分、関数を実行する
#- 1行目を左寄せ及び変数columns_0Xと一致した1行目の列名を色付けする
ChangeExcelColor_CLASS.color_columns(sheet_name_01, pink_columns['columns'],pink_columns['color']) #* classメソッド

ChangeExcelColor_CLASS.color_columns(sheet_name_01, orange_columns['columns'],orange_columns['color']) #* classメソッド

ChangeExcelColor_CLASS.color_columns(sheet_name_01, lightBlue_columns['columns'],lightBlue_columns['color']) #* classメソッド

ChangeExcelColor_CLASS.color_columns(sheet_name_01, blue_columns['columns'],blue_columns['color']) #* classメソッド

ChangeExcelColor_CLASS.color_columns(sheet_name_01, lightYellow_columns['columns'],lightYellow_columns['color']) #* classメソッド

ChangeExcelColor_CLASS.color_columns(sheet_name_01, yellow_columns['columns'],yellow_columns['color']) #* classメソッド

#- フィルター及び行の固定を適用 - 20220725
ChangeExcelColor_CLASS.filter_and_freeze_rows(sheet_name_01)

#- 上記の prefix に入る文字を引数として送る - 20220607
ChangeExcelColor_CLASS.save_workbook('colored') #* classメソッド
#^  ---------------------- 20220606 -------------------------

module_01.open_folder(result_file_location)

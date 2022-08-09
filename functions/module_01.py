import pandas as pd
import glob
import datetime
import re

#- 参照を2行目から最終行まで行う列を色付けするために導入 - 20220405
import openpyxl
from openpyxl.styles import PatternFill
#excel配置 https://www.shibutan-bloomers.com/python_libraly_openpyxl-5/2664/
from openpyxl.styles import Alignment    # Alignmentクラスをインポート
#- 20220725
from openpyxl.utils import get_column_letter, column_index_from_string
#- 20220607
import subprocess
import os
from pathlib import Path


t_delta = datetime.timedelta(hours=9)
JST = datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)
# print(repr(now))

today = now.strftime('%Y%m%d')

d = now.strftime('%Y%m%d%H%M%S')
# print(d)  # 20211104173728 秒まで

d = now.strftime('%Y%m%d%H%M')
# print(d)  # 202111041737 分まで

#- 関数: スクレイピングでダウンロードしたexcelファイル全部のブランド名をまとめて一つのファイル名にする関数 - 20220331 ebay用
def concat_file_names():

    file_name_container = []

    excel_files = glob.glob('/Users/aokihirotaka/Desktop/python_lesson_20220203/csv_folder/csv_for_concat/ebay/*.xlsx')

    supplier = excel_files[0].split('-')[1]
    supplier

    # https://tinyurl.com/y78xfprs split()でいらない単語を削除
    # https://tinyurl.com/y7r3f7kv 各単語のタイトルをの最初の文字を大文字化
    for i, file in enumerate(excel_files):
        file_name = file.split('-')[2].title()
        file_name_container.append(file_name)
        # print('\n',i,file_name)
        # print('\n',i,file_name_container)

    file_name_complete = '-'.join(file_name_container).replace(' ','')
    file_name_complete = f'final-{supplier}-{file_name_complete}-{today}.xlsx'
    print('\n','file_name_complete: ',file_name_complete)

    print('\n',file_name_complete)


    return file_name_complete


#- 関数: スクレイピングでダウンロードしたexcelファイル全部のブランド名をまとめて一つのファイル名にする関数 - 20220411 shopify用
class Concat_file_names:

    #! コンストラクタ不要
    #- そもそも__init__メソッドはオブジェクトを返せません。(註)=Noneは返すことができる
    # def __init__(self, file_loc):
    #     self.file_loc = file_loc
    #     file_location = self.file_loc
        # file_location = self.file_location

    #! クラス内のメソッドの引数には必ず self を入れる必要有り!! 20220412
    def concat_file_names(self, file_location, file_search_word):

        file_name_container = []
        # excel_files = glob.glob(f'{file_location}*.xlsx')
        # excel_files = glob.glob(f'{file_location}*2nd*.xlsx')
        #- 2ndSt以外から取得したファイルも取得できるように file_search_wordで検索できるようにした - 20220424
        excel_files = glob.glob(f'{file_location}*{file_search_word}*.xlsx')

        print('\n', 'excel_files: ', excel_files ,'\n')

        #- shopifyのファイルとそれ以外のファイルで場合分け - ファイル名の記入ルールが異なるため
        #- shopifyファイルの場合('shopify'という文字がファイル名に含まれる場合):
        if 'shopify' in excel_files[0]:
            supplier = excel_files[0].split('-')[1]

            # https://tinyurl.com/y78xfprs split()でいらない単語を削除
            # https://tinyurl.com/y7r3f7kv 各単語のタイトルをの最初の文字を大文字化

            for i, file in enumerate(excel_files):
                file_name = file.split('-')[2].title()
                file_name_container.append(file_name)
                print('\n',i,'file_name: ',file_name)
                print('\n',i,'file_name_container: ',file_name_container)

                #- ファイル名に性別とカテゴリーを追加する-20220502
                if i == len(excel_files)-1:
                    file_gender = file.split('-')[-2].title()
                    file_category = file.split('-')[-3].title()
                    print(file_gender)
                    print(file_category)
                    file_name_container.append(file_gender)
                    file_name_container.append(file_category)


        #- shopifyファイル以外の場合:
        else:
            #* 例:final-/Users/aokihirotaka/Desktop/python_lesson_20220203/csv_folder/csv_for_concat/ebay/trf-Jilsander-Fearofgod-20220428.xlsx
            #* の中の '-' で区切った最初の節[0]の文字の後ろから3文字を抜き出す = 仕入先名
            # https://tinyurl.com/y6x94mps
            supplier = excel_files[0].split('-')[0][-3:]
            print('supplier: ',supplier)

            for i, file in enumerate(excel_files):
                file_name = file.split('-')[1].title()
                file_name_container.append(file_name)
                print('\n',i,'file_name: ',file_name)
                print('\n',i,'file_name_container: ',file_name_container)

                #- ファイル名に性別とカテゴリーを追加する-20220502
                if i == len(excel_files)-1:
                    file_gender = file.split('-')[-2].title()
                    # file_category = file.split('-')[-3].title()
                    file_category = file.split('-')[-3]
                    print(file_gender)
                    print(file_category)
                    file_name_container.append(file_gender)
                    file_name_container.append(file_category)

        file_name_complete = '-'.join(file_name_container).replace(' ','')
        file_name_complete = f'final-{supplier}-{file_name_complete}-{today}.xlsx'
        print('\n','file_name_complete: ',file_name_complete)

        print('\n',file_name_complete)


        return file_name_complete

    # def add(a, b):
    #     return a + b

    # def sub(a, b):
    #     return a - b

#- 関数: 文字列が英数字かどうかを判断する(正規表現)
#- def check_alphabet(item_title_string): で使用する
#! https://vatchlog.com/python-alnum/
def checkAlnum(word):
    alnum = re.compile(r'^[a-zA-Z0-9]+$')
    result = alnum.match(word) is not None
    return result

#- 関数: リストの中の最初の2つの要素の文字列が、英数字かどうかをチェックし、英数字ならそのまま残し、そうでないなら削除した後に、残りのリストの要素をjoin()で一つの文字列 = item_title に直す

def check_alphabet(item_title_string):

    item_title_container = []

    word_list = item_title_string.split(' ')
    # print(word_list)

    for i, each_word in enumerate(word_list):
        # 英数字なら True、それ以外の日本語等は Falseを返す
        result = checkAlnum(each_word)
        if i <= 1:
            # print('def check_alphabet(item_title_string): - result: ',result,i)
            if result==True:
                item_title_container.append(each_word)
        else:
            item_title_container.append(each_word)

    item_title = ' '.join(item_title_container).strip()
    return item_title

#- 関数: 2ndStの商品からCategory Detailを取得する - バックパック、ポーチ等 - 20220407
#* shopify用
def get_type_detail(TYPE_DETAIL):

    #例 TYPE_DETAIL = 'clare v/トートバッグ/--'
    # => ['clare v', 'トートバッグ', '--']

    list_type_detail= TYPE_DETAIL.split('/') # / を区切りにリスト化

    # 必要に応じて増やしていく
    types = {'ウエストバッグ':'Waist bag','ショルダーバッグ':'Shoulder bag','トートバッグ':'Tote bag','バックパック':'Backpack','ビジネスバッグ':'Business bag','ボストンバッグ':'Boston bag','トラベルバッグ':'Travel bag','クラッチバッグ':'Clutch bag','ハンドバッグ':'Handbag','ポーチ':'Pouch','その他':'Others','ウエストバック':'Waist bag','ショルダーバック':'Shoulder bag','トートバック':'Tote bag','バックパック':'Backpack','ビジネスバック':'Business bag','ボストンバック':'Boston bag','トラベルバック':'Travel bag','クラッチバック':'Clutch bag','ハンドバック':'Handbag','リュック':'Backpack','ガーメントバッグ':'Garment bag','ガーメントバック':'Garment bag','ブリーフケース':'Business bag','トラベルキャリー':'Travel bag','セカンドバッグ':'Second bag','セカンドバック':'Second bag'}

    [{'abc':'1'}]

    matchedType = ''

    for i, type in enumerate(list_type_detail):
        for key, value in types.items():
            # if type == key:
            #例: キャンバストートバッグ <= type - 2ndSt商品パンくずリスト内
            #- 修正-完全一致ではなく、部分一致にも対応20220413
            if key in type:
                matchedType = value
                break

    print('\n', 'FROM MODULE - matchedType: ', matchedType ,'\n')
    return matchedType

#- 関数: pythonで自動取得した商品シートをconcatした後、replace.pyで特定のwordを置き換えた後のシートの特定列の1行目を色付けする関数
def change_excel_columns_color(OUTPUT_FILENAME):
    #エクセルファイル・シートを指定
    wb = openpyxl.load_workbook(f'/Users/aokihirotaka/Desktop/python_lesson_20220203/{OUTPUT_FILENAME}')
    ws_main = wb['main']
    ws_reference = wb['reference']

    # ws_combined = wb['main','reference']

    #探す文字列を指定
    ref_value_main = ["*Title",	"Ref-len()","C:Brand","ConditionDescription","X-condition_detail","Ref-condition-detail_translation","ref-url","costPrice + s/p","PicURL","MinimumBestOfferPrice"]

    ref_value_reference = ["URef-size_for_ConditionDescription","Ref-Brand","Ref-color-translation","Ref-condition_translation","Ref-condition-detail_translation","Ref-costPrice + s/p","Ref-features_translation","Ref-gender-translation","Ref-material_translation","Ref-model-translation","Ref-PicURL","Ref-size-cm_translation","Ref-size-inch_translation","Ref-title_category_combined","ref-title-completed","Ref-translation(Title)","Ref-URL","URef-title-completed"]

    #[My_Value]をループ
    for list in ref_value_main:

        #1行ずつループ
        for row in ws_main.iter_rows():

            #1セルずつループ
            for cell in row:

                #セルに [My_Value] の文字列が含まれていたら
                if list in str(cell.value):

                    #該当セルに色付け
                    cell.fill = PatternFill(fgColor='ff00ff',bgColor="ff00ff", fill_type = "solid")


    #[My_Value]をループ
    for list in ref_value_reference:

        #1行ずつループ
        for row in ws_reference.iter_rows():

            #1セルずつループ
            for cell in row:

                #セルに [My_Value] の文字列が含まれていたら
                if list in str(cell.value):

                    #該当セルに色付け
                    cell.fill = PatternFill(fgColor='ff00ff',bgColor="ff00ff", fill_type = "solid")

    #別名で保存
    wb.save(f'/Users/aokihirotaka/Desktop/python_lesson_20220203/000-vec-colored-{OUTPUT_FILENAME}')

    print('\n', 'File-name: ', f'000-vec-colored-{OUTPUT_FILENAME}'  ,'\n')


#- class クラス: shopifyで使用 - pythonで自動取得した商品シートをconcatした後、replace.pyで特定のwordを置き換えた後のシートの特定列の1行目を色付けする関数
# 20220412 キノコ-ド クラス    https://www.youtube.com/watch?v=F5guF1y7G48
class ChangeExcelColor:
    def change_excel_columns_color(self, load_workbook_location, OUTPUT_FILENAME, result_file_location, sheet_name01, sheet_name02):

        #- エクセルファイル・シートを指定
        # self.wb = openpyxl.load_workbook(f'/Users/aokihirotaka/Desktop/python_lesson_20220203/{OUTPUT_FILENAME}')
        #-load_workbook_location引数でファイルの指定を柔軟にした - 20220510
        self.wb = openpyxl.load_workbook(f'{load_workbook_location}/{OUTPUT_FILENAME}')
        # self.ws_main = self.wb['main']
        # self.ws_reference = self.wb['reference']
        self.ws_main = self.wb[sheet_name01]
        self.ws_reference = self.wb[sheet_name02]

        #! 下記だと左揃えにならない for分が必要-20220510
        # https://gammasoft.jp/blog/text-center-alignment-with-openpyxl/

        # self.ws_main.alignment = Alignment(horizontal='left')   # 左揃え
        # self.ws_reference.alignment = Alignment(horizontal='left')   # 左揃え

        # ws_combined = wb['main','reference']

        #[My_Value]をループ
        #- 下記アトリビュート self.ref_value_mainは、replace-shopify.pyの中で作成している
        #- => よってこのclass内にはself.ref_value_mainの記述がなくても使えている
        # ピンク色
        for list in self.ref_value_main:

            #1行ずつループ
            for row in self.ws_main.iter_rows():

                #1セルずつループ
                for cell in row:


                    #! 全てのセルを網羅するとファイル作成に時間がかかるため1行目のみ適用 => 断然早くなった! - 2022041
                    # https://qiita.com/github-nakasho/items/facb51ead1cf1d2dd232
                    if cell.row == 1:

                        # 全てのcellを左揃えにする - 20220412
                        cell.alignment = Alignment(horizontal='left')

                        #セルに [My_Value] の文字列が含まれていたら
                        # if list in str(cell.value):
                        #* 文字完全一致した場合のみ色付け
                        if list == str(cell.value):

                            #該当セルに色付け
                            cell.fill = PatternFill(fgColor='ff00ff',bgColor="ff00ff", fill_type = "solid")

        #!!!!!!!!!!!!!!!!!!

        # 青色
        for list in self.ref_value_main_blue:

            #1行ずつループ
            for row in self.ws_main.iter_rows():

                #1セルずつループ
                for cell in row:

                    #! 全てのセルを網羅するとファイル作成に時間がかかるため1行目のみ適用 => 断然早くなった! - 2022041
                    # https://qiita.com/github-nakasho/items/facb51ead1cf1d2dd232
                    if cell.row == 1:

                        # 全てのcellを左揃えにする - 20220412
                        cell.alignment = Alignment(horizontal='left')

                        #セルに [My_Value] の文字列が含まれていたら
                        # if list in str(cell.value):
                        #* 文字完全一致した場合のみ色付け
                        if list == str(cell.value):

                            #該当セルに色付け
                            cell.fill = PatternFill(fgColor='00FFFF',bgColor="00FFFF", fill_type = "solid")


        #[My_Value]をループ
        #- 上記アトリビュート self.ref_value_mainと同様
        # ピンク色
        for list in self.ref_value_reference:

            #1行ずつループ
            for row in self.ws_reference.iter_rows():

                #1セルずつループ
                for cell in row:

                    #! 全てのセルを網羅するとファイル作成に時間がかかるため1行目のみ適用 => 断然早くなった! - 20220412
                    # https://qiita.com/github-nakasho/items/facb51ead1cf1d2dd232
                    if cell.row == 1:

                        # 全てのcellを左揃えにする - 20220412
                        cell.alignment = Alignment(horizontal='left')

                        #セルに [My_Value] の文字列が含まれていたら
                        # if list in str(cell.value):
                        #* 文字完全一致した場合のみ色付け
                        if list == str(cell.value):

                            #該当セルに色付け
                            cell.fill = PatternFill(fgColor='ff00ff',bgColor="ff00ff", fill_type = "solid")


        #[My_Value]をループ
        #- 上記アトリビュート self.ref_value_mainと同様
        #- 全ての行(セル)を埋める必要のある列を色付けする
        # 水色
        for list in self.ref_value_reference_dragAll:

            #1行ずつループ
            for row in self.ws_reference.iter_rows():

                #1セルずつループ
                for cell in row:

                    #! 全てのセルを網羅するとファイル作成に時間がかかるため1行目のみ適用 => 断然早くなった! - 20220412
                    # https://qiita.com/github-nakasho/items/facb51ead1cf1d2dd232
                    if cell.row == 1:

                        # 全てのcellを左揃えにする - 20220412
                        cell.alignment = Alignment(horizontal='left')

                        #セルに [My_Value] の文字列が含まれていたら
                        # if list in str(cell.value):
                        #* 文字完全一致した場合のみ色付け
                        if list == str(cell.value):

                            #該当セルに色付け
                            cell.fill = PatternFill(fgColor='00FFFF',bgColor='00FFFF', fill_type = 'solid')

                            break

        #- シートの移動 - 20220527 https://tinyurl.com/2mmtkgrm
        # 「offset=-1」であれば左へ一つ分移動します。
        try:
            self.wb.move_sheet('reference', offset=-1)
        except Exception as e:
            print('\n','例外発生 :',e)

        #別名で保存
        # self.wb.save(f'/Users/aokihirotaka/Desktop/python_lesson_20220203/000-vec-colored-{OUTPUT_FILENAME}')

        self.wb.save(f'{result_file_location}/colored-{OUTPUT_FILENAME}')

        # print('\n', 'File-name: ', f'000-vec-colored-{OUTPUT_FILENAME}'  ,'\n')
        print('\n', '色付け後保存場所: ', f'{result_file_location}/colored-{OUTPUT_FILENAME}'  ,'\n')

#- 20220607 - 改善版
class ChangeExcelColorRevised:
    def __init__(self, load_workbook_location, OUTPUT_FILENAME, result_file_location):
        self.load_workbook_location = load_workbook_location
        self.OUTPUT_FILENAME = OUTPUT_FILENAME
        self.result_file_location = result_file_location
        #- 読み込むエクセルファイルを指定
        self.wb = openpyxl.load_workbook(f'{self.load_workbook_location}/{self.OUTPUT_FILENAME}')

    #- selected_columnsは、色を変えたい対象の行のリスト
    #- colorは塗りつぶしの色 'ff00ff' <= ピンク
    def color_columns(self, sheet_name, color_columns, color):
        try:
            for list in color_columns:

                #1行ずつループ
                # for row in self.ws_sheet01.iter_rows():
                # for row in self.wb[self.ws_sheet01].iter_rows():
                for row in self.wb[sheet_name].iter_rows():

                    #1セルずつループ
                    for cell in row:
                        # print('cell: ',cell)

                        #! 全てのセルを網羅するとファイル作成に時間がかかるため1行目のみ適用 => 断然早くなった! - 2022041
                        # https://qiita.com/github-nakasho/items/facb51ead1cf1d2dd232
                        if cell.row == 1:

                            # 全てのcellを左揃えにする - 20220412
                            cell.alignment = Alignment(horizontal='left')

                            #セルに [My_Value] の文字列が含まれていたら
                            # if list in str(cell.value):
                            #* 文字完全一致した場合のみ色付け
                            if list == str(cell.value):

                                #該当セルに色付け
                                # cell.fill = PatternFill(fgColor='ff00ff',bgColor="ff00ff", fill_type = "solid")
                                cell.fill = PatternFill(fgColor=color,bgColor=color, fill_type = 'solid')

                                # print('\n', 'list: ', list ,'\n')
                                # print('\n', 'cell.value: ', cell.value ,'\n')
                                #- breakを入れることでifで見つけた後もループし続ける必要がないため、コードが早く終る
                                break

        except Exception as e:
            print('例外 from color_first_row(): ',e)

    #- シートの移動 - 20220527 https://tinyurl.com/2mmtkgrm
    # 「offset=-1」であれば左へ一つ分移動します。
    def move_sheet(self,sheet_name):
        self.wb.move_sheet(sheet_name, offset=1)

    #- フィルター及び行の固定を適用 - 20220725
    # https://tinyurl.com/28lj6umk
    def filter_and_freeze_rows(self,sheet_name):
        #- フィルター適用
        self.wb[sheet_name].auto_filter.ref = get_column_letter(1) + str(1) + ':' + get_column_letter(self.wb[sheet_name].max_column) + str(self.wb[sheet_name].max_row)

        #- 行の固定適用 https://tinyurl.com/25qbdld5
        self.wb[sheet_name].freeze_panes = 'A3'

    #- シートの保存 - prefixにはファイル名の前につけたい言葉を入れる
    def save_workbook(self,prefix=''):
        # self.wb.save(f'{self.result_file_location}/watchCount-{self.OUTPUT_FILENAME}')
        self.wb.save(f'{self.result_file_location}/{prefix}-{self.OUTPUT_FILENAME}')

        print('\n', '色付け後保存場所: ', f'{self.result_file_location}/colored-{self.OUTPUT_FILENAME}'  ,'\n')

        print(f'★★★ファイル名: colored-{self.OUTPUT_FILENAME}★★★','\n')

    def save_workbook_original_title(self):
        # self.wb.save(f'{self.result_file_location}/watchCount-{self.OUTPUT_FILENAME}')
        self.wb.save(f'{self.result_file_location}/{self.OUTPUT_FILENAME}')

        print('\n', 'save_workbook_original_title() - 色付け後保存場所: ', f'{self.result_file_location}/colored-{self.OUTPUT_FILENAME}'  ,'\n')

        print(f'★★★ファイル名: {self.OUTPUT_FILENAME}★★★','\n')



#- Pythonのsubprocessモジュールでいつも使うファイル・フォルダを自動で開くプログラム-20220607
def open_folder(folder_dir):
    print('folder_dir: ', folder_dir)

    #- プログラム終了後、フォルダーを開く  https://tinyurl.com/25ojjpdl
    if 'C:' in str(folder_dir):
        #* windowsでsubprocess.Popen()を使ってフォルダを開く場合、
        #* / ではなく \ である必要がある! - 20220615
        # folder_dir  = folder_dir.replace('/','\\')
        folder_dir  = str(folder_dir).replace('/',os.sep)
        print("フォルダーを開くコード: ", "subprocess.Popen(['explorer',f'{folder_dir}']},shell=True)")
        subprocess.Popen(['explorer',folder_dir],shell=True)
    else:
        print("フォルダーを開くコード: ", "subprocess.call(['open', f'{folder_dir}'])")
        subprocess.call(['open', f'{folder_dir}'])

#- WindowsとMacで互換性のあるpathを作成するクラス - 20220608
class Control_path:
    def __init__(self):
        #- module_01.pyのディレクトリではなく、呼び出し元のファイルのcwdが入手できる
        self.cwd = Path.cwd()
        self.cwd_str = str(self.cwd)

    #- windows変更 \を/に変更した- 20220529
    #^ current pathを取得 - ここを元にpathを指定していく
    #! パスを取得する関数を利用した場合、パスの区切り文字は「\」バックスラッシュになります。
    #! .replace(os.sep,'/')で対応する!!
    #- cwd 現在path
    print('os.sep: ', os.sep)

    #- 相対パスで取得したいファイルのディレクトリのを区切り文字を / に変更し、かつパスオブジェクトから文字列に変換して取得する関数
    def neutralize_path_str(self,path):
        #- パスを繋げ、パスオブジェクトから文字列に変換する
        new_path_str = str(self.cwd / path)
        #- パス同士をつなげ完成したファイルのディレクトリの区切り文字を / に変更
        return new_path_str.replace(os.sep,'/')

    #- 相対パスで取得したいファイルディレクトリをオブジェクトで取得する
    def neutralize_path_obj(self, path):
        #- パスを繋げ、パスオブジェクトを作成
        new_path_obj = self.cwd / path
        #^ TypeError: replace() takes 2 positional arguments but 3 were given
        #! 不可 オブジェクトにreplace()を使おうとすると上記のエラーが発生する
        # return new_path.replace(os.sep,'/')
        return new_path_obj

    #- ChromeDriverの場所を指定する
    def locate_chromeDriver(self,path):
        chromeDriverLocation = str(list(self.neutralize_path_obj(path).glob('*'))[0]).replace(os.sep,'/')
        return chromeDriverLocation

    #- 親ディレクトリを取得 - 20220608
    #! ここで取得するよりも呼び出し元のファイルで作成したほうが早いし、よさそう
    def get_parent_dir(self, path):
        return path.parent

    #- 特定のpathをstrで取得する
    def get_specific_dir_str(self, basic_path, added_path):
        specific_dir_str = str(basic_path / added_path)
        return specific_dir_str

    #- 特定のpathをオブジェクトで取得する
    def get_specific_dir_obj(self, basic_path, added_path):
        specific_dir_obj = basic_path / added_path
        return specific_dir_obj
